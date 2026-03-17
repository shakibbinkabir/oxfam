"""
Seed climate indicators from the Tech Team Excel spreadsheet.

Parses Tech_Team_Climate_Risk_Calculation.xlsx, handles merged cells
by carrying forward the last non-null Component and Subcategory values,
and upserts all 67 indicators into the climate_indicators table.

Usage:
    python -m app.scripts.seed_indicators
    python -m app.scripts.seed_indicators --file path/to/excel.xlsx
"""

import argparse
import asyncio
import logging
import sys
from pathlib import Path

from sqlalchemy import text

from app.database import async_session

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")
logger = logging.getLogger(__name__)


def parse_excel(filepath: str) -> list[dict]:
    """Parse the climate indicator Excel file with merged cell handling."""
    try:
        import openpyxl
    except ImportError:
        logger.error("openpyxl is required: pip install openpyxl")
        sys.exit(1)

    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active

    indicators = []
    current_component = None
    current_subcategory = None
    seen_codes = set()

    for row_idx in range(2, ws.max_row + 1):
        component_val = ws.cell(row=row_idx, column=1).value
        subcategory_val = ws.cell(row=row_idx, column=2).value
        indicator_name = ws.cell(row=row_idx, column=3).value
        code_val = ws.cell(row=row_idx, column=4).value
        unit = ws.cell(row=row_idx, column=5).value
        source = ws.cell(row=row_idx, column=6).value
        gis_attr = ws.cell(row=row_idx, column=7).value

        # Skip empty rows
        if not indicator_name:
            continue

        # Carry forward merged cell values
        if component_val:
            current_component = str(component_val).strip()
        if subcategory_val:
            current_subcategory = str(subcategory_val).strip()

        # Reset subcategory when component changes (new component without explicit subcategory)
        if component_val and not subcategory_val:
            current_subcategory = None

        indicator_name = str(indicator_name).strip()

        # Build a unique code: use the explicit code column if present,
        # otherwise use gis_attribute_id, and make it unique by prefixing
        # with component abbreviation if there's a collision
        base_code = str(code_val).strip() if code_val else (str(gis_attr).strip() if gis_attr else None)
        if not base_code:
            base_code = indicator_name.lower().replace(" ", "_")[:30]

        # Ensure uniqueness by prefixing with component abbreviation
        code = base_code
        if code in seen_codes:
            prefix_map = {
                "Hazard": "haz",
                "Socioeconomic": "soc",
                "Environmental": "env",
                "Infrastructural": "inf",
            }
            prefix = prefix_map.get(current_component, current_component[:3].lower())
            subcat_short = ""
            if current_subcategory:
                subcat_map = {
                    "Exposure": "exp",
                    "Sensitivity": "sen",
                    "Adaptive Capacity": "ac",
                    "Adaptive Capacity ": "ac",
                    "Hazard": "",
                }
                subcat_short = subcat_map.get(current_subcategory.strip(), current_subcategory[:3].lower())

            if subcat_short:
                code = f"{prefix}_{subcat_short}_{base_code}"
            else:
                code = f"{prefix}_{base_code}"

            # If still not unique, add a numeric suffix
            if code in seen_codes:
                counter = 2
                while f"{code}_{counter}" in seen_codes:
                    counter += 1
                code = f"{code}_{counter}"

        seen_codes.add(code)

        # For Hazard component, subcategory comes from the Excel merged cell ("Hazard")
        subcat = current_subcategory
        if current_component == "Hazard" and not subcat:
            subcat = "Hazard"

        indicators.append({
            "component": current_component,
            "subcategory": subcat,
            "indicator_name": indicator_name,
            "code": code,
            "unit": str(unit).strip() if unit else None,
            "source": str(source).strip() if source else None,
            "gis_attribute_id": str(gis_attr).strip() if gis_attr else None,
        })

    return indicators


async def seed_indicators(indicators: list[dict]):
    """Upsert indicators into the database, creating units and sources as needed."""
    async with async_session() as session:
        count = 0
        for ind in indicators:
            unit_id = None
            source_id = None

            # Resolve unit
            if ind.get("unit"):
                result = await session.execute(
                    text("SELECT id FROM units WHERE name = :name"),
                    {"name": ind["unit"]},
                )
                row = result.one_or_none()
                if row:
                    unit_id = row.id
                else:
                    await session.execute(
                        text("INSERT INTO units (name) VALUES (:name) ON CONFLICT (name) DO NOTHING"),
                        {"name": ind["unit"]},
                    )
                    result = await session.execute(
                        text("SELECT id FROM units WHERE name = :name"),
                        {"name": ind["unit"]},
                    )
                    unit_id = result.one().id

            # Resolve source
            if ind.get("source"):
                result = await session.execute(
                    text("SELECT id FROM sources WHERE name = :name"),
                    {"name": ind["source"]},
                )
                row = result.one_or_none()
                if row:
                    source_id = row.id
                else:
                    await session.execute(
                        text("INSERT INTO sources (name) VALUES (:name) ON CONFLICT (name) DO NOTHING"),
                        {"name": ind["source"]},
                    )
                    result = await session.execute(
                        text("SELECT id FROM sources WHERE name = :name"),
                        {"name": ind["source"]},
                    )
                    source_id = result.one().id

            sql = text("""
                INSERT INTO climate_indicators
                    (component, subcategory, indicator_name, code, unit_id, source_id, gis_attribute_id)
                VALUES
                    (:component, :subcategory, :indicator_name, :code, :unit_id, :source_id, :gis_attribute_id)
                ON CONFLICT (code) DO UPDATE SET
                    component = EXCLUDED.component,
                    subcategory = EXCLUDED.subcategory,
                    indicator_name = EXCLUDED.indicator_name,
                    unit_id = EXCLUDED.unit_id,
                    source_id = EXCLUDED.source_id,
                    gis_attribute_id = EXCLUDED.gis_attribute_id,
                    updated_at = NOW()
            """)
            await session.execute(sql, {
                "component": ind["component"],
                "subcategory": ind["subcategory"],
                "indicator_name": ind["indicator_name"],
                "code": ind["code"],
                "unit_id": unit_id,
                "source_id": source_id,
                "gis_attribute_id": ind.get("gis_attribute_id"),
            })
            count += 1

        await session.commit()
        logger.info(f"Seeded {count} climate indicators")

        # Verify by component
        result = await session.execute(
            text("SELECT component, COUNT(*) as cnt FROM climate_indicators GROUP BY component ORDER BY component")
        )
        for row in result.all():
            logger.info(f"  {row.component}: {row.cnt} indicators")


def main():
    parser = argparse.ArgumentParser(description="Seed climate indicators from Excel")
    parser.add_argument(
        "--file",
        default="./Tech Team_Climate Risk_Calculation.xlsx",
        help="Path to the Excel file",
    )
    args = parser.parse_args()

    filepath = Path(args.file)
    if not filepath.exists():
        logger.error(f"Excel file not found: {filepath}")
        sys.exit(1)

    indicators = parse_excel(str(filepath))
    logger.info(f"Parsed {len(indicators)} indicators from {filepath.name}")

    for ind in indicators:
        logger.info(f"  [{ind['component']}] [{ind['subcategory'] or '-'}] {ind['code']}: {ind['indicator_name']}")

    asyncio.run(seed_indicators(indicators))


if __name__ == "__main__":
    main()
